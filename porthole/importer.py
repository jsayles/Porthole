from openpyxl import load_workbook

from django.conf import settings
from django.utils.timezone import localtime, now

from porthole.models import Location, Switch, VLAN, Port


LOCATIONS = "Locations"
SWITCHES = "Switches"
VLANS = "VLANs"
PORTS = "Ports"


######################################################################
# Utilities
######################################################################


def sheet_to_dict(sheet):
    keys = [sheet.cell(1, col_index).value.lower() for col_index in range(1, sheet.max_column + 1)]
    dict_list = []
    for row_index in range(2, sheet.max_row + 1):
        row_dict = {}
        for col_index in range(1, sheet.max_column + 1):
            key = keys[col_index - 1]
            row_dict[key] = sheet.cell(row_index, col_index).value
        dict_list.append(row_dict)
    return dict_list


######################################################################
# Data Importer
######################################################################


class Importer(object):

    def __init__(self, file_name):
        self.workbook = load_workbook(filename = file_name)
        for sheet in [LOCATIONS, SWITCHES, VLANS, PORTS]:
            if not sheet in self.workbook.sheetnames:
                raise Exception("'%s' Sheet Not Found!" % sheet)

    def import_data(self):
        self.import_locations()
        self.import_switches()
        self.import_vlans()
        self.import_ports()

    def import_locations(self):
        sheet = self.workbook[LOCATIONS]
        locations = sheet_to_dict(sheet)
        for row in locations:
            number = str(row['number'])
            location = Location.objects.filter(number=number).first()
            if not location:
                # Create a new location
                location = Location(number=number, name=row['name'], floor=int(number[0]))
            else:
                # Update the name on an existing location
                location.name = row['name']
            location.save()

    def import_switches(self):
        sheet = self.workbook[SWITCHES]
        switches = sheet_to_dict(sheet)
        for row in switches:
            label = row['label']
            switch = Switch.objects.filter(label=label).first()
            if not switch:
                location = Location.objects.filter(number=row['location']).first()
                if not location:
                    raise Exception("Location '%s' not found!" % row['location'])
                switch = Switch(label=label, location=location)
            switch.make = row['make']
            switch.model = row['model']
            switch.port_count = row['port_count']
            switch.save()

    def import_vlans(self):
        sheet = self.workbook[VLANS]
        for row in sheet_to_dict(sheet):
            tag = row['tag']
            vlan = VLAN.objects.filter(tag=tag).first()
            if not vlan:
                vlan = VLAN(tag=tag)
            vlan.name = row['name']
            vlan.description = row['description']
            vlan.ip_range = row['ip_range']
            vlan.save()

    def import_ports(self):
        sheet = self.workbook[PORTS]
        for row in sheet_to_dict(sheet):
            # Find the location
            location = Location.objects.filter(number=row['location']).first()
            if not location:
                raise Exception("Location '%s' not found!" % row['location'])

            # Find the VLAN
            vlan = VLAN.objects.filter(tag=row['vlan']).first()
            if not vlan:
                raise Exception("VLAN '%s' not found!" % row['vlan'])

            # Find the switch
            switch = Switch.objects.filter(label=row['switch']).first()

            # We may have 1 or 2 labels on a single line
            # NOTE:  This won't work if there is no space in the label name
            port_labels = []
            if row['port'].startswith("AB"):
                port_prefix, port_number = row['port'].split(" ")
                port_labels.append("A " + port_number)
                port_labels.append("B " + port_number)
            else:
                port_labels.append(row['port'])

            # Create or update the ports
            for label in port_labels:
                port = Port.objects.filter(label=label).first()
                if not port:
                    port = Port(label=label)
                port.location = location
                port.vlan = vlan
                port.switch = switch
                port.switch_port = row['switch port']
                port.save()
